import os
from openpyxl import Workbook, load_workbook

# Ruta del archivo Excel
EXCEL_PATH = 'usuarios.xlsx'

# Si el archivo no existe, lo creamos e insertamos encabezados
if not os.path.exists(EXCEL_PATH):
    wb = Workbook()
    ws = wb.active
    ws.append(['Nombre', 'Apellido', 'Correo'])
    wb.save(EXCEL_PATH)
    print(f"Archivo creado con encabezados: {EXCEL_PATH}")

# Cargamos (o reusamos) el libro de Excel
wb = load_workbook(EXCEL_PATH)
ws = wb.active

# Cantidad de registros que el usuario desea ingresar
try:
    n = int(input('¿Cuántos registros deseas ingresar? '))
except ValueError:
    print("Por favor ingresa un número válido.")
    exit(1)

for i in range(1, n + 1):
    print(f"\n--- Registro {i} ---")
    nombre = input('Nombre: ').strip()
    apellido = input('Apellido: ').strip()
    correo = input('Correo: ').strip()
    ws.append([nombre, apellido, correo])

# Guardamos los cambios en el archivo Excel
wb.save(EXCEL_PATH)
print(f"\n¡Se han agregado {n} registros correctamente a '{EXCEL_PATH}'!")
